from pathlib import Path
import openpyxl

contract = Path('functions/geminiModerationContract.js')
s = contract.read_text()
covered_needle = "  '- underwear, lingerie, swimwear, a thong, or a string when intimate areas remain covered.',\n  '',\n  'Sexual context can still make a clothed image adult-only:',"
covered_repl = "  '- underwear, lingerie, swimwear, a thong, or a string when intimate areas remain covered.',\n  '',\n  'adultDecision describes nudity / sexual explicitness only; adult-only sexual context is represented separately by triggers.',\n  '- Covered lingerie/underwear remains adultDecision=\\\"none\\\" when breasts, nipples, genitals and buttocks are not exposed and the visible body/clothing/crop does not create the appearance of nudity.',\n  '- BDSM equipment, rope, restraints, fetish styling, erotic posing or sexual suggestion do NOT by themselves count as implied nudity. They may require adultEroticSuggestive/kinkBdsm triggers while adultDecision stays \\\"none\\\".',\n  '- Do not infer nudity merely because a clothed BDSM/kink image is sexual or fetishistic.',\n  '',\n  'Sexual context can still make a clothed image adult-only:',"
explicit_needle = "  '- \"explicit\": a clear sexual act is visible, such as penetration, oral sex, masturbation, genital stimulation, or another unambiguous sex act.',"
explicit_repl = explicit_needle + "\n  '- Visible genitals alone are NOT a sexual act. A close-up of exposed genitals, including visible natural moisture or a droplet/fluid, stays adultDecision=\\\"borderline\\\" unless the image also visibly shows touching/stimulation, penetration, oral contact, masturbation or another unambiguous sex act.'" + "\n  '- Do not infer masturbation, stimulation or penetration from arousal, wetness, bodily fluid, pose, framing or a genital close-up alone. The act itself must be visibly evidenced.'"
if covered_needle not in s:
    raise SystemExit('covered prompt insertion point not found')
if explicit_needle not in s:
    raise SystemExit('explicit prompt insertion point not found')
contract.write_text(s.replace(covered_needle, covered_repl, 1).replace(explicit_needle, explicit_repl, 1))

test_path = Path('functions/test/geminiModerationContract.test.js')
t = test_path.read_text()
nudity_assert = "  assert.match(prompt, /underwear, lingerie, swimwear, a thong, or a string/i);"
erotic_assert = "  assert.match(prompt, /penetration, oral sex, masturbation/i);"
if nudity_assert not in t or erotic_assert not in t:
    raise SystemExit('contract test insertion point not found')
t = t.replace(
    nudity_assert,
    nudity_assert + "\n  assert.match(prompt, /BDSM equipment, rope, restraints.*do NOT by themselves count as implied nudity/i);\n  assert.match(prompt, /adult-only sexual context is represented separately by triggers/i);",
    1,
)
t = t.replace(
    erotic_assert,
    erotic_assert + "\n  assert.match(prompt, /Visible genitals alone are NOT a sexual act/i);\n  assert.match(prompt, /Do not infer masturbation, stimulation or penetration from arousal, wetness, bodily fluid/i);",
    1,
)
test_path.write_text(t)

xlsx_path = Path('testing/moderation_goldens_v1/testcases_reviewed_v3.xlsx')
wb = openpyxl.load_workbook(xlsx_path)
found_adult = False
found_borderline = False
for ws in wb.worksheets:
    headers = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value is not None}
    if 'case_id' not in headers:
        continue
    for r in range(2, ws.max_row + 1):
        case = ws.cell(r, headers['case_id']).value
        if case == 'SAFE_01':
            values = {
                'case_id': 'ADULT_BDSM_01',
                'filename': 'images/adult/ADULT_BDSM_01.jpg',
                'category': 'adult',
                'safety_tag_inputs': '18+ BDSM / kink',
                'expected_content_truth': 'covered BDSM/kink without nudity or explicit sexual act',
                'expected_ai_result': 'adult_context',
                'expected_policy_result': 'adult',
                'expected_final_result': 'adult',
                'expected_ui_result': 'visible only with active 18+ entitlement',
                'notes': 'Policy v2 correction after non-production live classifier inspection: this legacy SAFE_01 image contains covered BDSM/kink context. It is adult via erotic/kink triggers, but adultDecision must remain none because no nudity or explicit sexual act is visible.',
            }
            for key, value in values.items():
                if key in headers:
                    ws.cell(r, headers[key]).value = value
            found_adult = True
        elif case == 'BORDERLINE_01':
            values = {
                'expected_ai_result': 'adult_nudity',
                'expected_policy_result': 'adult',
                'expected_final_result': 'adult',
                'expected_ui_result': 'visible only with active 18+ entitlement',
                'notes': 'Policy v2 correction: visible genitalia without a visible sexual act is adult nudity, not explicit sexual content and not automatically review. Human review is reserved for genuine safety or explicit-act uncertainty.',
            }
            for key, value in values.items():
                if key in headers:
                    ws.cell(r, headers[key]).value = value
            found_borderline = True
if not found_adult or not found_borderline:
    raise SystemExit(f'xlsx rows not found: adult={found_adult} borderline={found_borderline}')
wb.save(xlsx_path)
